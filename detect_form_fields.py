import openpyxl
from openpyxl.utils import get_column_letter

# File path
excel_file = r'C:\Users\banu.parasuraman\Downloads\SailPoint_Onboarding_Application_Questionnaire_v2.xlsx'

# Load the workbook
wb = openpyxl.load_workbook(excel_file)

print("="*80)
print("DETAILED SHEET ANALYSIS - Finding Form Fields")
print("="*80)

# Focus on the main sheets
main_sheets = ['Application General Information', 'Application On-boarding Form', 'Process type ', 'Environment']

for sheet_name in main_sheets:
    if sheet_name not in wb.sheetnames:
        continue
        
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"SHEET: {sheet_name}")
    print(f"{'='*80}\n")
    
    # Find all non-empty cells and their values
    form_fields = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column), start=1):
        for col_idx, cell in enumerate(row, start=1):
            if cell.value and str(cell.value).strip():
                # Check if next cell is empty (likely a field to fill)
                next_cell = ws.cell(row=row_idx, column=col_idx+1)
                is_label = next_cell.value is None or str(next_cell.value).strip() == ""
                
                # Store potential form fields
                if is_label and len(str(cell.value)) < 200:  # Likely a field label
                    form_fields.append({
                        'row': row_idx,
                        'col': col_idx,
                        'cell_ref': f"{get_column_letter(col_idx)}{row_idx}",
                        'label': str(cell.value).strip(),
                        'value_cell': f"{get_column_letter(col_idx+1)}{row_idx}"
                    })
    
    # Print form fields
    if form_fields:
        print(f"Found {len(form_fields)} potential form fields:\n")
        for idx, field in enumerate(form_fields[:30], 1):  # Show first 30
            print(f"{idx:2d}. [{field['cell_ref']}] {field['label'][:60]}")
            print(f"    → Value goes in: {field['value_cell']}\n")

wb.close()
