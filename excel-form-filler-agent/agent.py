import json
import requests
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime
from typing import Optional
from google.adk.agents.llm_agent import Agent

MCP_QUERY_URL = 'http://34.9.116.130:3000/mcp/query'

def safe_set_cell(ws, cell_ref, value):
    """Safely set cell value, handling merged cells."""
    try:
        cell = ws[cell_ref]
        # For merged cells, openpyxl only allows writing to the top-left cell
        # Check if this cell is part of a merged range
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                # Get the top-left cell of the merged range
                top_left_cell = ws.cell(merged_range.min_row, merged_range.min_col)
                top_left_cell.value = value
                return
        # Not a merged cell, set directly
        cell.value = value
    except Exception as e:
        # If there's any issue, try to set the value directly
        try:
            ws[cell_ref].value = value
        except:
            pass  # Skip if unable to write

# Get the project root directory (parent of excel-form-filler-agent)
PROJECT_ROOT = Path(__file__).parent.parent
DEFAULT_TEMPLATE = str(PROJECT_ROOT / 'SailPoint_Onboarding_Application_Questionnaire_v2.xlsx')

def get_sailpoint_data_from_mcp() -> dict:
    """Fetches SailPoint onboarding data from the MCP server."""
    try:
        response = requests.post(MCP_QUERY_URL, json={'query': 'get_user_data'})
        response.raise_for_status()
        
        users = response.json()
        
        # Extract schema information
        if users:
            schema = {
                'identityAttribute': 'userId',
                'displayAttribute': 'email',
                'attributes': list(users[0].keys())
            }
            
            # Extract unique roles
            all_roles = [role for user in users for role in user.get('roles', [])]
            entitlements = list(set(all_roles))
            
            return {
                "status": "success",
                "user_count": len(users),
                "users": users,
                "schema": schema,
                "entitlements": entitlements
            }
        else:
            return {
                "status": "success",
                "user_count": 0,
                "users": [],
                "schema": {},
                "entitlements": []
            }
    
    except requests.exceptions.RequestException as e:
        return {
            "status": "error",
            "message": f"Failed to fetch user data: {str(e)}"
        }

def fill_excel_form(template_path: Optional[str] = None, output_path: Optional[str] = None) -> dict:
    """
    Fills the SailPoint onboarding Excel form with data from MCP server.
    
    Args:
        template_path: Path to the Excel template file (optional, uses default if not provided)
        output_path: Path where the filled Excel file should be saved
    
    Returns:
        Dictionary with status and details of the operation
    """
    try:
        # Use default template if none provided
        if template_path is None:
            template_path = DEFAULT_TEMPLATE
        
        # Convert to Path object for easier handling
        template_path = Path(template_path)
        
        # Check if template exists
        if not template_path.exists():
            return {
                "status": "error",
                "message": f"Template file not found: {template_path}"
            }
        
        # Get data from MCP server
        data = get_sailpoint_data_from_mcp()
        
        if data["status"] == "error":
            return data
        
        # Load the Excel template
        wb = openpyxl.load_workbook(str(template_path))
        
        # Fill Application General Information sheet
        if 'Application General Information' in wb.sheetnames:
            ws = wb['Application General Information']
            
            # Application Name
            safe_set_cell(ws, 'C12', "MongoDB Authorization App")
            
            # Application Description  
            safe_set_cell(ws, 'C19', "MongoDB-based application managing user authorization, roles, and permissions for internal systems.")
            
            # Submitted By
            safe_set_cell(ws, 'D21', "Automated Agent")
            
            # Business Owner
            safe_set_cell(ws, 'C25', "IT Security Team")
            
            # Technical Owner
            safe_set_cell(ws, 'C26', "Database Administrator")
            
            # Lead Technical Contact
            safe_set_cell(ws, 'C27', "MongoDB Team")
            
            # Internally developed or procured
            safe_set_cell(ws, 'C29', "Internally Developed")
            
            # Environments
            safe_set_cell(ws, 'C30', "DEV, UAT, PROD")
            
            # Domain
            safe_set_cell(ws, 'C31', "CWS")
            
            # Business Objectives
            safe_set_cell(ws, 'C32', "Centralized user access management and role-based authorization")
            
            # SOW Required
            safe_set_cell(ws, 'C33', "No")
            
            # Total active users
            safe_set_cell(ws, 'C34', str(data.get('user_count', 0)))
            
            # Uses Entra/AD
            safe_set_cell(ws, 'C35', "No - uses MongoDB for authentication")
            
            # Current provisioning process
            safe_set_cell(ws, 'C36', "Manual - Database updates")
            
            # Account creation process
            safe_set_cell(ws, 'C37', "Direct MongoDB document insertion with role assignment")
            
            # Account types
            safe_set_cell(ws, 'C38', "Employee, Contractor")
            
            # Total roles
            safe_set_cell(ws, 'C39', f"{len(data.get('entitlements', []))} roles: {', '.join(data.get('entitlements', []))}")
            
            # Multiple roles
            safe_set_cell(ws, 'C40', "Yes")
            
            # Elevated privileges
            safe_set_cell(ws, 'C41', "Admin role provides elevated access to system configuration")
            
            # RBAC
            safe_set_cell(ws, 'C42', "Yes - role-based access controls implemented")
            
            # Super admin
            safe_set_cell(ws, 'C43', "Yes - 'admin' role has full system access")
            
            # SOD Policies
            safe_set_cell(ws, 'C44', "Yes - admin and user roles have separation")
        
        # Fill Application On-boarding Form sheet
        if 'Application On-boarding Form' in wb.sheetnames:
            ws = wb['Application On-boarding Form']
            
            # Case sensitive
            safe_set_cell(ws, 'C13', "Yes")
            
            # Authorized to change
            safe_set_cell(ws, 'C14', "Database Administrator")
            
            # Disabled accounts
            safe_set_cell(ws, 'C15', "Yes - status='inactive'")
            
            # Dormant accounts
            safe_set_cell(ws, 'C16', "Yes - via status field")
            
            # Service accounts
            safe_set_cell(ws, 'C17', "Yes - MongoDB connection credentials")
            
            # Password rotation
            safe_set_cell(ws, 'C18', "Quarterly")
            
            # Business Owner
            safe_set_cell(ws, 'C20', "IT Security Team")
            
            # Attribute Name
            safe_set_cell(ws, 'C21', "userId, firstName, lastName, email, status, roles")
        
        # Fill Environment sheet
        if 'Environment' in wb.sheetnames:
            ws = wb['Environment']
            
            # Production Details
            safe_set_cell(ws, 'G13', "34.172.211.78")  # Hostname
            safe_set_cell(ws, 'G14', "27017")  # Port
            safe_set_cell(ws, 'G15', "app_auth")  # Database Name
            safe_set_cell(ws, 'G16', "sailpoint_readonly")  # Username (placeholder)
            safe_set_cell(ws, 'G17', "[Stored in Secrets Manager]")  # Password
            safe_set_cell(ws, 'G18', "mongodb://34.172.211.78:27017/app_auth")  # JDBC/Connection URL
            
            # API Details
            safe_set_cell(ws, 'G23', "http://34.9.116.130:3000")  # Base URL (MCP Server)
        
        # Fill Process type sheet
        if 'Process type ' in wb.sheetnames:
            ws = wb['Process type ']
            
            # Create account
            safe_set_cell(ws, 'B3', "Yes")
            safe_set_cell(ws, 'C3', "Create user document in MongoDB with required attributes and roles")
            
            # Modify account
            safe_set_cell(ws, 'B4', "Yes")
            safe_set_cell(ws, 'C4', "Update user document fields including roles array")
            
            # Disable account
            safe_set_cell(ws, 'B5', "Yes")
            safe_set_cell(ws, 'C5', "Set status field to 'inactive'")
            
            # Delete account
            safe_set_cell(ws, 'B6', "Yes")
            safe_set_cell(ws, 'C6', "Remove user document from collection")
            
            # Required attributes
            safe_set_cell(ws, 'B7', "userId, firstName, lastName, email, status, roles[]")
        
        # Fill Roles sheet
        if 'Roles' in wb.sheetnames:
            ws = wb['Roles']
            
            # Fill production roles
            roles_data = data.get('entitlements', [])
            start_row = 2
            for idx, role in enumerate(roles_data, start=1):
                safe_set_cell(ws, f'G{start_row + idx}', idx)  # S.no
                safe_set_cell(ws, f'H{start_row + idx}', role)  # Role Name
                safe_set_cell(ws, f'I{start_row + idx}', f"{role} permissions")  # Entitlement
                safe_set_cell(ws, f'J{start_row + idx}', f"Standard {role} role")  # Description
        
        # Generate output path if not provided
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_dir = Path(__file__).parent
            output_path = output_dir / f"SailPoint_Onboarding_Filled_{timestamp}.xlsx"
        
        # Save the filled workbook
        wb.save(str(output_path))
        wb.close()
        
        return {
            "status": "success",
            "message": f"Excel form filled successfully and saved to: {output_path}",
            "output_file": str(output_path),
            "total_accounts": data["user_count"],
            "entitlements": data["entitlements"],
            "filled_sheets": ["Application General Information", "Application On-boarding Form", "Environment", "Process type ", "Roles"]
        }
    
    except Exception as e:
        return {
            "status": "error",
            "message": f"Error filling Excel form: {str(e)}"
        }

def read_excel_form(file_path: str) -> dict:
    """
    Reads the filled SailPoint onboarding Excel form and returns the data.
    
    Args:
        file_path: Path to the Excel file to read
    
    Returns:
        Dictionary containing all the form data
    """
    try:
        if not Path(file_path).exists():
            return {
                "status": "error",
                "message": f"File not found: {file_path}"
            }
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        form_data = {
            "application_details": {
                "application_name": ws['B4'].value,
                "application_owner": ws['B5'].value,
                "description": ws['B6'].value,
                "application_type": ws['B7'].value
            },
            "connection_details": {
                "connector_type": ws['B9'].value,
                "host": ws['B10'].value,
                "port": ws['B11'].value,
                "database": ws['B12'].value,
                "jdbc_url": ws['B13'].value,
                "authentication": ws['B14'].value
            },
            "schema_mapping": {
                "identity_attribute": ws['B16'].value,
                "display_attribute": ws['B17'].value,
                "account_attributes": ws['B18'].value,
                "entitlement_attribute": ws['B19'].value
            },
            "entitlements": {
                "discovered_roles": ws['B21'].value,
                "entitlement_type": ws['B22'].value
            },
            "account_correlation": {
                "correlation_rule": ws['B24'].value,
                "correlation_attribute": ws['B25'].value
            },
            "provisioning_policy": {
                "create_account": ws['B27'].value,
                "update_account": ws['B28'].value,
                "delete_account": ws['B29'].value,
                "manage_entitlements": ws['B30'].value
            },
            "aggregation_info": {
                "total_accounts": ws['B32'].value,
                "sample_account_1": ws['B33'].value,
                "sample_account_2": ws['B34'].value,
                "sample_account_3": ws['B35'].value
            }
        }
        
        return {
            "status": "success",
            "form_data": form_data
        }
    
    except Exception as e:
        return {
            "status": "error",
            "message": f"Error reading Excel form: {str(e)}"
        }

# Create the Excel Form Filler Agent
root_agent = Agent(
    model='gemini-2.0-flash-exp',
    name='excel_form_filler_agent',
    description="An agent that fills SailPoint onboarding Excel forms with data from MongoDB via MCP server.",
    instruction="""You are an Excel form automation specialist for SailPoint application onboarding.

When a user asks you to fill the Excel form:
1. Call 'fill_excel_form' tool WITHOUT providing template_path parameter (leave it empty/null) - it will use the default template
2. The tool automatically fetches data from MCP/MongoDB and fills all fields in 5 sheets: Application General Information, Application On-boarding Form, Environment, Process type, and Roles
3. Provide a clear summary of what was filled and where the file was saved

IMPORTANT: Do NOT specify a template_path when calling fill_excel_form. The default template is automatically used.

When a user asks to read a filled form:
1. Use 'read_excel_form' tool with the file path
2. Present the data in a clear, organized format

When a user asks for raw MongoDB data:
1. Use 'get_sailpoint_data_from_mcp' tool

The template is automatically located in the project root directory.
Output files are saved in the excel-form-filler-agent directory with a timestamp.""",
    tools=[fill_excel_form, read_excel_form, get_sailpoint_data_from_mcp],
)
