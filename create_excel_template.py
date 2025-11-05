import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "SailPoint Onboarding Form"

# Define styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=12)
section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
section_font = Font(bold=True, size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Set column widths
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 50

# Title
ws.merge_cells('A1:B1')
ws['A1'] = "SailPoint Application Onboarding Form"
ws['A1'].font = Font(bold=True, size=14)
ws['A1'].alignment = Alignment(horizontal='center')
ws['A1'].fill = header_fill
ws['A1'].font = header_font

row = 3

# Section 1: Application Details
ws[f'A{row}'] = "1. Application Details"
ws[f'A{row}'].font = section_font
ws[f'A{row}'].fill = section_fill
ws.merge_cells(f'A{row}:B{row}')
row += 1

fields = [
    ("Application Name", ""),
    ("Application Owner", ""),
    ("Description", ""),
    ("Application Type", "")
]

for field, value in fields:
    ws[f'A{row}'] = field
    ws[f'B{row}'] = value
    ws[f'A{row}'].border = border
    ws[f'B{row}'].border = border
    row += 1

row += 1

# Section 2: Connection Details
ws[f'A{row}'] = "2. Connection Details"
ws[f'A{row}'].font = section_font
ws[f'A{row}'].fill = section_fill
ws.merge_cells(f'A{row}:B{row}')
row += 1

fields = [
    ("Connector Type", ""),
    ("Host", ""),
    ("Port", ""),
    ("Database", ""),
    ("JDBC URL", ""),
    ("Authentication Method", "")
]

for field, value in fields:
    ws[f'A{row}'] = field
    ws[f'B{row}'] = value
    ws[f'A{row}'].border = border
    ws[f'B{row}'].border = border
    row += 1

row += 1

# Section 3: Schema Mapping
ws[f'A{row}'] = "3. Schema Mapping"
ws[f'A{row}'].font = section_font
ws[f'A{row}'].fill = section_fill
ws.merge_cells(f'A{row}:B{row}')
row += 1

fields = [
    ("Identity Attribute", ""),
    ("Display Attribute", ""),
    ("Account Attributes", ""),
    ("Entitlement Attribute", "")
]

for field, value in fields:
    ws[f'A{row}'] = field
    ws[f'B{row}'] = value
    ws[f'A{row}'].border = border
    ws[f'B{row}'].border = border
    row += 1

row += 1

# Section 4: Entitlements
ws[f'A{row}'] = "4. Entitlements"
ws[f'A{row}'].font = section_font
ws[f'A{row}'].fill = section_fill
ws.merge_cells(f'A{row}:B{row}')
row += 1

fields = [
    ("Discovered Roles", ""),
    ("Entitlement Type", "")
]

for field, value in fields:
    ws[f'A{row}'] = field
    ws[f'B{row}'] = value
    ws[f'A{row}'].border = border
    ws[f'B{row}'].border = border
    row += 1

row += 1

# Section 5: Account Correlation
ws[f'A{row}'] = "5. Account Correlation"
ws[f'A{row}'].font = section_font
ws[f'A{row}'].fill = section_fill
ws.merge_cells(f'A{row}:B{row}')
row += 1

fields = [
    ("Correlation Rule", ""),
    ("Correlation Attribute", "")
]

for field, value in fields:
    ws[f'A{row}'] = field
    ws[f'B{row}'] = value
    ws[f'A{row}'].border = border
    ws[f'B{row}'].border = border
    row += 1

row += 1

# Section 6: Provisioning Policy
ws[f'A{row}'] = "6. Provisioning Policy"
ws[f'A{row}'].font = section_font
ws[f'A{row}'].fill = section_fill
ws.merge_cells(f'A{row}:B{row}')
row += 1

fields = [
    ("Create Account", ""),
    ("Update Account", ""),
    ("Delete Account", ""),
    ("Manage Entitlements", "")
]

for field, value in fields:
    ws[f'A{row}'] = field
    ws[f'B{row}'] = value
    ws[f'A{row}'].border = border
    ws[f'B{row}'].border = border
    row += 1

row += 1

# Section 7: Aggregation Info
ws[f'A{row}'] = "7. Aggregation Information"
ws[f'A{row}'].font = section_font
ws[f'A{row}'].fill = section_fill
ws.merge_cells(f'A{row}:B{row}')
row += 1

fields = [
    ("Total Accounts", ""),
    ("Sample Account 1", ""),
    ("Sample Account 2", ""),
    ("Sample Account 3", "")
]

for field, value in fields:
    ws[f'A{row}'] = field
    ws[f'B{row}'] = value
    ws[f'A{row}'].border = border
    ws[f'B{row}'].border = border
    row += 1

# Save the workbook
wb.save('SailPoint_Onboarding_Template.xlsx')
print("Excel template created successfully: SailPoint_Onboarding_Template.xlsx")
